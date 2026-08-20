#!/usr/bin/env python3
"""
External Integration Test Case Normalizer

外部持込の結合試験項目表を読み込み、意味を変更せずCanonical JSONへ正規化する。

Supported input:
    - .xlsx
    - .csv
    - .json

Default input directory:
    external-tests/integration-test/

Default output:
    reports/integration-test/external-test-cases.normalized.json

Exit codes:
    0: Success
    1: Validation / processing failure
    2: External test input required

Dependencies:
    openpyxl  (.xlsx利用時のみ)
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import sys
from pathlib import Path
from typing import Any


SUPPORTED_SUFFIXES = {".xlsx", ".csv", ".json"}

REQUIRED_COLUMNS = [
    "case_id",
    "title",
    "requirement_id",
    "test_category",
    "criterion",
    "steps",
    "expected_result",
    "execution_type",
]

OPTIONAL_COLUMNS = [
    "related_adr",
    "precondition",
    "input",
    "notes",
]

ALL_COLUMNS = REQUIRED_COLUMNS + OPTIONAL_COLUMNS

PROTECTED_SEMANTIC_FIELDS = [
    "case_id",
    "title",
    "requirement_id",
    "related_adr",
    "test_category",
    "criterion",
    "precondition",
    "input",
    "steps",
    "expected_result",
    "execution_type",
    "notes",
]

ALLOWED_EXECUTION_TYPES = {
    "AUTOMATABLE",
    "NOT_AUTOMATABLE",
}


class InputRequiredError(RuntimeError):
    """外部試験項目表が未配置で、ユーザー確認が必要な場合。"""


def normalize_text(value: Any) -> str:
    """セル値を比較可能な文字列へ変換する。意味を変える加工はしない。"""
    if value is None:
        return ""

    text = str(value)
    text = text.replace("\r\n", "\n").replace("\r", "\n")
    return text.strip()


def nullable_text(value: Any) -> str | None:
    text = normalize_text(value)
    return text if text else None


def semantic_payload(row: dict[str, Any]) -> dict[str, str]:
    """
    外部ケースの意味を構成する保護対象項目だけを取り出す。
    HashはこのPayloadに対して計算する。
    """
    return {
        field: normalize_text(row.get(field))
        for field in PROTECTED_SEMANTIC_FIELDS
    }


def semantic_hash(row: dict[str, Any]) -> str:
    payload = semantic_payload(row)

    encoded = json.dumps(
        payload,
        ensure_ascii=False,
        sort_keys=True,
        separators=(",", ":"),
    ).encode("utf-8")

    digest = hashlib.sha256(encoded).hexdigest()
    return f"sha256:{digest}"


def discover_criteria(criteria_dir: Path) -> set[str]:
    if not criteria_dir.exists():
        return set()

    result: set[str] = set()

    for path in criteria_dir.glob("*.criterion.md"):
        if path.is_file():
            result.add(path.name.removesuffix(".criterion.md"))

    return result


def resolve_input(
    explicit_input: Path | None,
    input_dir: Path,
    allow_none: bool,
) -> Path | None:

    if explicit_input is not None:
        if not explicit_input.exists():
            raise FileNotFoundError(
                f"External test case file not found: {explicit_input}"
            )

        if explicit_input.suffix.lower() not in SUPPORTED_SUFFIXES:
            raise ValueError(
                f"Unsupported external test case format: {explicit_input.suffix}"
            )

        return explicit_input

    if not input_dir.exists():
        if allow_none:
            return None

        raise InputRequiredError(
            f"External test case directory does not exist: {input_dir}"
        )

    candidates = sorted(
        path
        for path in input_dir.iterdir()
        if path.is_file()
        and path.suffix.lower() in SUPPORTED_SUFFIXES
        and not path.name.startswith("~$")
    )

    if not candidates:
        if allow_none:
            return None

        raise InputRequiredError(
            f"No external integration test case file found in: {input_dir}"
        )

    if len(candidates) > 1:
        names = ", ".join(path.name for path in candidates)
        raise ValueError(
            "Multiple external test case files found. "
            "Specify one with --input. "
            f"Candidates: {names}"
        )

    return candidates[0]


def load_xlsx(path: Path, sheet_name: str) -> list[dict[str, Any]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError(
            "openpyxl is required to read .xlsx files. "
            "Install it with: pip install openpyxl"
        ) from exc

    workbook = load_workbook(
        filename=path,
        read_only=True,
        data_only=True,
    )

    if sheet_name not in workbook.sheetnames:
        raise ValueError(
            f"Worksheet '{sheet_name}' not found in {path}. "
            f"Available: {', '.join(workbook.sheetnames)}"
        )

    sheet = workbook[sheet_name]

    rows = sheet.iter_rows(values_only=True)

    try:
        raw_header = next(rows)
    except StopIteration:
        return []

    headers = [
        normalize_text(value)
        for value in raw_header
    ]

    data: list[dict[str, Any]] = []

    for values in rows:
        row = {
            headers[index]: values[index]
            for index in range(min(len(headers), len(values)))
            if headers[index]
        }

        if any(normalize_text(value) for value in row.values()):
            data.append(row)

    return data


def load_csv(path: Path) -> list[dict[str, Any]]:
    with path.open(
        "r",
        encoding="utf-8-sig",
        newline="",
    ) as file:
        reader = csv.DictReader(file)
        return [dict(row) for row in reader]


def load_json(path: Path) -> list[dict[str, Any]]:
    with path.open("r", encoding="utf-8") as file:
        data = json.load(file)

    if isinstance(data, dict):
        cases = data.get("cases")
    else:
        cases = data

    if not isinstance(cases, list):
        raise ValueError(
            "JSON input must be an array or an object containing a 'cases' array."
        )

    result: list[dict[str, Any]] = []

    for index, item in enumerate(cases, start=1):
        if not isinstance(item, dict):
            raise ValueError(
                f"JSON case #{index} must be an object."
            )
        result.append(item)

    return result


def load_rows(path: Path, sheet_name: str) -> list[dict[str, Any]]:
    suffix = path.suffix.lower()

    if suffix == ".xlsx":
        return load_xlsx(path, sheet_name)

    if suffix == ".csv":
        return load_csv(path)

    if suffix == ".json":
        return load_json(path)

    raise ValueError(f"Unsupported format: {suffix}")


def validate_columns(rows: list[dict[str, Any]]) -> list[str]:
    errors: list[str] = []

    if not rows:
        return ["External test case file contains no test cases."]

    columns = set()

    for row in rows:
        columns.update(row.keys())

    for column in REQUIRED_COLUMNS:
        if column not in columns:
            errors.append(
                f"Required column missing: {column}"
            )

    return errors


def validate_and_normalize_rows(
    rows: list[dict[str, Any]],
    source_file: Path,
    criteria: set[str],
) -> tuple[list[dict[str, Any]], list[dict[str, Any]], list[str]]:

    errors: list[str] = []
    normalized_cases: list[dict[str, Any]] = []
    integrity: list[dict[str, Any]] = []

    seen_case_ids: set[str] = set()

    for row_number, row in enumerate(rows, start=2):
        source_case_id = normalize_text(row.get("case_id"))

        label = (
            source_case_id
            if source_case_id
            else f"row-{row_number}"
        )

        for field in REQUIRED_COLUMNS:
            if not normalize_text(row.get(field)):
                errors.append(
                    f"{label}: required field '{field}' is empty."
                )

        if not source_case_id:
            continue

        if source_case_id in seen_case_ids:
            errors.append(
                f"Duplicate external case_id: {source_case_id}"
            )
            continue

        seen_case_ids.add(source_case_id)

        execution_type = normalize_text(
            row.get("execution_type")
        ).upper()

        if (
            execution_type
            and execution_type not in ALLOWED_EXECUTION_TYPES
        ):
            errors.append(
                f"{source_case_id}: invalid execution_type "
                f"'{execution_type}'."
            )

        criterion = normalize_text(row.get("criterion"))

        if criteria and criterion and criterion not in criteria:
            errors.append(
                f"{source_case_id}: unknown criterion '{criterion}'. "
                "Add the matching *.criterion.md or correct the test item."
            )

        source_hash = semantic_hash(row)

        canonical: dict[str, Any] = {
            "case_id": source_case_id,
            "origin": "EXTERNAL",
            "source_case_id": source_case_id,
            "source_file": str(source_file),
            "title": normalize_text(row.get("title")),
            "requirement_id": normalize_text(row.get("requirement_id")),
            "related_adr": nullable_text(row.get("related_adr")),
            "test_category": normalize_text(row.get("test_category")),
            "criterion": criterion,
            "precondition": nullable_text(row.get("precondition")),
            "input": nullable_text(row.get("input")),
            "steps": normalize_text(row.get("steps")),
            "expected_result": normalize_text(row.get("expected_result")),
            "execution_type": execution_type,
            "notes": nullable_text(row.get("notes")),
            "coverage_key": None
        }

        normalized_hash = semantic_hash(canonical)

        normalized_cases.append(canonical)

        integrity.append(
            {
                "case_id": source_case_id,
                "source_case_id": source_case_id,
                "source_file": str(source_file),
                "source_semantic_hash": source_hash,
                "normalized_semantic_hash": normalized_hash,
                "integrity_status": (
                    "PASS"
                    if source_hash == normalized_hash
                    else "FAIL"
                ),
            }
        )

        if source_hash != normalized_hash:
            errors.append(
                f"{source_case_id}: semantic content changed during normalization."
            )

    return normalized_cases, integrity, errors


def build_empty_output() -> dict[str, Any]:
    return {
        "external_cases_provided": False,
        "source_file": None,
        "cases": [],
        "external_case_integrity": [],
    }


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Normalize external integration test cases "
            "without changing their semantic content."
        )
    )

    parser.add_argument(
        "--input",
        type=Path,
        default=None,
        help="Explicit .xlsx/.csv/.json external test case file.",
    )

    parser.add_argument(
        "--input-dir",
        type=Path,
        default=Path("external-tests/integration-test"),
    )

    parser.add_argument(
        "--sheet",
        default="IntegrationTestCases",
        help="Worksheet name for .xlsx input.",
    )

    parser.add_argument(
        "--criteria-dir",
        type=Path,
        default=Path(
            ".github/skills/integration-test/criteria"
        ),
    )

    parser.add_argument(
        "--output",
        type=Path,
        default=Path(
            "reports/integration-test/"
            "external-test-cases.normalized.json"
        ),
    )

    parser.add_argument(
        "--allow-none",
        action="store_true",
        help=(
            "Proceed without external cases. "
            "Use only after the user explicitly confirms "
            "that no external cases will be provided."
        ),
    )

    return parser.parse_args()


def main() -> int:
    args = parse_args()

    try:
        source_file = resolve_input(
            explicit_input=args.input,
            input_dir=args.input_dir,
            allow_none=args.allow_none,
        )

        args.output.parent.mkdir(
            parents=True,
            exist_ok=True,
        )

        if source_file is None:
            args.output.write_text(
                json.dumps(
                    build_empty_output(),
                    ensure_ascii=False,
                    indent=2,
                ),
                encoding="utf-8",
            )

            print(
                "[PASS] No external test cases were provided "
                "and the user explicitly allowed continuation."
            )
            print(f"Output: {args.output}")
            return 0

        rows = load_rows(
            source_file,
            args.sheet,
        )

        errors = validate_columns(rows)

        criteria = discover_criteria(
            args.criteria_dir
        )

        (
            normalized_cases,
            integrity,
            row_errors,
        ) = validate_and_normalize_rows(
            rows=rows,
            source_file=source_file,
            criteria=criteria,
        )

        errors.extend(row_errors)

        if errors:
            print(
                "========================================"
            )
            print(
                " External Integration Test Normalization"
            )
            print(
                "========================================"
            )
            print()
            print("[FAIL] External test case validation failed.")
            print()

            for index, error in enumerate(errors, start=1):
                print(f"{index}. {error}")

            return 1

        output = {
            "external_cases_provided": True,
            "source_file": str(source_file),
            "case_count": len(normalized_cases),
            "cases": normalized_cases,
            "external_case_integrity": integrity,
        }

        args.output.write_text(
            json.dumps(
                output,
                ensure_ascii=False,
                indent=2,
            ),
            encoding="utf-8",
        )

        print(
            f"[PASS] Normalized {len(normalized_cases)} "
            "external integration test case(s)."
        )
        print(f"Output: {args.output}")
        return 0

    except InputRequiredError as error:
        print(
            "========================================"
        )
        print(
            " External Integration Test Input Required"
        )
        print(
            "========================================"
        )
        print()
        print(f"[INPUT_REQUIRED] {error}")
        print()
        print(
            "Place a completed integration test case file under:"
        )
        print("  external-tests/integration-test/")
        print()
        print(
            "Template:"
        )
        print(
            "  .github/skills/integration-test/templates/"
            "integration-test-case-template.xlsx"
        )
        print()
        print(
            "If no external cases will be used, "
            "the user must explicitly confirm continuation "
            "without external cases."
        )
        return 2

    except (
        FileNotFoundError,
        RuntimeError,
        ValueError,
        json.JSONDecodeError,
    ) as error:
        print(f"[FAIL] {error}")
        return 1


if __name__ == "__main__":
    sys.exit(main())
